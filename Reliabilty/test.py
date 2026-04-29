import os
import pm4py
from openpyxl import Workbook

# ======================================================
# 1️⃣ 路径设置
# ======================================================
log_dir = r"F:/402/模型挖掘算法可靠性评价方法/结果/Synthetic log4/相似度_models_single_reference_IMi"

standard_log_path = r"F:/402/模型挖掘算法可靠性评价方法/结果/Synthetic log4/Synthetic log 4.xes"

output_excel = r"F:/402/模型挖掘算法可靠性评价方法/结果/Synthetic log4/structural_similarity_results_IMi.xlsx"

# ======================================================
# 2️⃣ 标准模型（只计算一次）
# ======================================================
print("读取标准日志...")
standard_log = pm4py.read_xes(standard_log_path)

print("挖掘标准模型...")
std_net, std_im, std_fm = pm4py.discover_petri_net_inductive(standard_log)

# ======================================================
# 3️⃣ Excel初始化
# ======================================================
wb = Workbook()
ws = wb.active

# ⭐ 第一列标题
ws.cell(row=1, column=1).value = "日志"
ws.cell(row=2, column=1).value = "相似度"

col_index = 2   # 从第二列开始写数据

# ======================================================
# 4️⃣ 遍历模型文件
# ======================================================
for file in os.listdir(log_dir):

    if file.endswith(".pnml"):

        model_path = os.path.join(log_dir, file)
        print(f"\n===== 处理 {file} =====")

        try:
            net, im, fm = pm4py.read_pnml(model_path)

            sim_value = pm4py.structural_similarity(
                net, im, fm,
                std_net, std_im, std_fm
            )

            print("结构相似度:", sim_value)

        except Exception as e:
            print("失败:", e)
            sim_value = "-"

        # ⭐ 写入Excel
        ws.cell(row=1, column=col_index).value = file
        ws.cell(row=2, column=col_index).value = sim_value

        col_index += 1

# ======================================================
# 5️⃣ 保存
# ======================================================
wb.save(output_excel)
print("\n所有结果已保存至:", output_excel)