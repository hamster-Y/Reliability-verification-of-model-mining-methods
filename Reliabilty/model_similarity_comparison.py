# import pm4py
#
# # pn,b,c=pm4py.read_pnml("F:/402/模型挖掘算法可靠性评价方法/结果/Synthetic log1/incremental_models_alpha/L4.pnml")
# # pn2,b2,c2=pm4py.read_pnml("F:/402/模型挖掘算法可靠性评价方法/结果/Synthetic log1/incremental_models_alpha/L8.pnml")
# # print(pm4py.check_is_workflow_net(pn))
# # print(pm4py.check_soundness(pn2,b2,c2))
# # print(pm4py.structural_similarity(pn,b,c,pn2,b2,c2))
#
# # l=pm4py.read_xes("F:/402/模型挖掘算法可靠性评价方法/结果/Synthetic log2/Synthetic log 2.xes")
# # a,b,c=pm4py.discover_petri_net_inductive(l)
# # pm4py.write_pnml(a,b,c,"F:/402/模型挖掘算法可靠性评价方法/结果/Synthetic log2/Synthetic log 2.pnml")
#
# import os
# import re
# import pm4py
# # from pm4py.objects.petri_net.importer import importer as pnml_importer
# # from pm4py.algo.analysis.woflan import algorithm as woflan
# # from pm4py.algo.evaluation.similarity import algorithm as sim
#
# from openpyxl import Workbook
#
#
# # ================== 配置路径 ==================
#
# # 存放多个 pnml 的文件夹
# PNML_DIR = r"F:/402/模型挖掘算法可靠性评价方法/结果/Synthetic log2/incremental_models_alpha"
#
# # 参考模型
# # REFERENCE_PNML = r"F:/402/模型挖掘算法可靠性评价方法/结果/Synthetic log2/Synthetic log 2.pnml"
#
# # 参考日志
# REFERENCE_LOG = r"F:/402/模型挖掘算法可靠性评价方法/结果/Synthetic log2/Synthetic log 2.xes"
#
# # 输出 Excel
# OUTPUT_XLSX = r"F:/402/模型挖掘算法可靠性评价方法/结果/Synthetic log2/similarity_result.xlsx"
#
#
# # ================== 判断是否为 WF-net ==================
#
# def is_wf_net(pn, im, fm):
#     """
#     使用 Woflan 判断 WF-net + soundness 前提
#     """
#     # try:
#     #     result = pm4py.check_is_workflow_net(pn)
#     #     # PM4Py 返回 dict
#     #     if isinstance(result, dict):
#     #         return result.get("is_workflow_net", False)
#
#     return pm4py.check_is_workflow_net(pn)
#
#     # except Exception as e:
#     #     print("WF-net check error:", e)
#     #     return False
#
#
# # ================== 主流程 ==================
# def extract_number(name):
#     nums = re.findall(r"\d+", name)
#     return int(nums[0]) if nums else 0
#
# def main():
#
#     # 1. 读取参考模型
#     # ref_pn, ref_im, ref_fm = pm4py.read_pnml(REFERENCE_PNML)
#     ref_log=pm4py.read_xes(REFERENCE_LOG)
#
#     print("Reference model loaded.")
#
#
#     # 2. 收集结果
#     file_names = []
#     similarities = []
#     fitnesss=[]
#     precisions=[]
#     f_measures=[]
#
#
#     # 3. 遍历文件夹
#     for file in sorted(os.listdir(PNML_DIR), key=extract_number):
#
#
#         if not file.lower().endswith(".pnml"):
#             continue
#
#         file_path = os.path.join(PNML_DIR, file)
#
#         print(f"Processing: {file}")
#
#         try:
#             # 读取模型
#             pn, im, fm = pm4py.read_pnml(file_path)
#
#             # 判断 WF-net
#             # if not is_wf_net(pn, im, fm):
#             #     print("  -> Not WF-net, skipped.")
#
#             # file_names.append(file)
#             # similarities.append(None)
#             # continue
#
#             fitness = pm4py.fitness_alignments(ref_log, pn, im, fm)["log_fitness"]
#             precision = pm4py.precision_alignments(ref_log, pn, im, fm)
#             # else:
#             #     fitness=0
#             #     precision = 0
#             f_measure = 0 if fitness + precision == 0 else 2 * fitness * precision / (fitness + precision)
#             # 计算行为轮廓相似度（推荐）
#             # sim_value=pm4py.structural_similarity(pn, im, fm, ref_pn, ref_im, ref_fm,)
#             # sim_value = sim.apply(
#             #     pn, im, fm,
#             #     ref_pn, ref_im, ref_fm,
#             #     variant=sim.Variants.BEHAVIORAL_PROFILE
#             # )
#
#             # print("  -> Similarity:", sim_value)
#
#             base_name = os.path.splitext(file)[0]
#             file_names.append(base_name)
#             fitnesss.append(fitness)
#             precisions.append(precision)
#             f_measures.append(f_measure)
#             # similarities.append(sim_value)
#
#
#         except Exception as e:
#
#             print("  -> Error:", e)
#
#             base_name = os.path.splitext(file)[0]
#             file_names.append(base_name)
#
#             similarities.append(None)
#
#
#     # 4. 写 Excel
#     write_excel(file_names, similarities)
#
#     print("Finished. Result saved to:", OUTPUT_XLSX)
#
#
# # ================== 写 Excel ==================
#
# def write_excel(names, sims):
#
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Similarity"
#
#     # 第一行：文件名
#     for col, name in enumerate(names, start=1):
#         ws.cell(row=1, column=col, value=name)
#
#     # 第二行：相似度
#     for col, fitnesss in enumerate(sims, start=1):
#         ws.cell(row=2, column=col, value=fitnesss)
#
#     for col, precisions in enumerate(sims, start=1):
#         ws.cell(row=3, column=col, value=precisions)
#
#     for col, f_measures in enumerate(sims, start=1):
#         ws.cell(row=4, column=col, value=f_measures)
#
#     wb.save(OUTPUT_XLSX)
#
#
# # ================== 入口 ==================
#
# if __name__ == "__main__":
#     main()
#
# import pm4py
# l=pm4py.read_xes("F:/402/模型挖掘算法可靠性评价方法/结果/Synthetic log2/Synthetic log 2.xes")
# a,b,c=pm4py.read_pnml("F:/402/模型挖掘算法可靠性评价方法/结果/Synthetic log1/incremental_models_alpha/L8.pnml")
# # print(pm4py.fitness_alignments(l,a,b,c))
# fitness = pm4py.fitness_alignments(l,a,b,c)["log_fitness"]
# precision = pm4py.precision_alignments(l,a,b,c)
# # else:
# #     fitness=0
# #     precision = 0
# f_measure = 0 if fitness + precision == 0 else 2 * fitness * precision / (fitness + precision)
# print(fitness,precision,f_measure)

import os
import re
import pm4py
from openpyxl import Workbook


# ================== 配置路径 ==================

# 存放多个 pnml 的文件夹
PNML_DIR = r"F:/402/模型挖掘算法可靠性评价方法/结果/Synthetic log5/incremental_models_IM"

# 参考模型
# REFERENCE_PNML = r"F:/402/模型挖掘算法可靠性评价方法/结果/Synthetic log2/Synthetic log 2.pnml"

# 参考日志
REFERENCE_XES = r"F:/402/模型挖掘算法可靠性评价方法/结果/Synthetic log5/Synthetic log 5.xes"

# 输出 Excel
OUTPUT_XLSX = r"F:/402/模型挖掘算法可靠性评价方法/结果/Synthetic log5/similarity_result_IM.xlsx"


# ================== 排序用 ==================

def extract_number(name):
    nums = re.findall(r"\d+", name)
    return int(nums[0]) if nums else 0


# ================== 主流程 ==================

def main():

    # 1. 读取参考模型 + 日志
    # ref_pn, ref_im, ref_fm = pm4py.read_pnml(REFERENCE_PNML)
    ref_log = pm4py.read_xes(REFERENCE_XES)

    print("Reference model and log loaded.")


    # 2. 收集结果
    file_names = []
    fitness_list = []
    precision_list = []
    fmeasure_list = []


    # 3. 遍历模型
    for file in sorted(os.listdir(PNML_DIR), key=extract_number):

        if not file.lower().endswith(".pnml"):
            continue

        file_path = os.path.join(PNML_DIR, file)

        print(f"Processing: {file}")

        base_name = os.path.splitext(file)[0]

        try:
            # 读取模型
            pn, im, fm = pm4py.read_pnml(file_path)

            # Fitness
            fitness = pm4py.fitness_alignments(
                ref_log, pn, im, fm
            )["log_fitness"]

            # Precision
            # precision = pm4py.precision_alignments(
            #     ref_log, pn, im, fm
            # )
            precision = pm4py.precision_token_based_replay( ref_log, pn, im, fm)

            # F-measure
            if fitness + precision == 0:
                f_measure = 0
            else:
                f_measure = 2 * fitness * precision / (fitness + precision)

            print(f"  -> fitness={fitness:.4f}, precision={precision:.4f}, F={f_measure:.4f}")

        except Exception as e:

            print("  -> Error:", e)

            fitness = None
            precision = None
            f_measure = None


        # 统一 append（保证长度一致）
        file_names.append(base_name)
        fitness_list.append(fitness)
        precision_list.append(precision)
        fmeasure_list.append(f_measure)


    # 4. 写 Excel
    write_excel(
        file_names,
        fitness_list,
        precision_list,
        fmeasure_list
    )

    print("Finished. Result saved to:", OUTPUT_XLSX)


# ================== 写 Excel ==================

def write_excel(names, fitnesss, precisions, f_measures):

    wb = Workbook()
    ws = wb.active
    ws.title = "Evaluation"

    # 第一行：文件名
    for col, name in enumerate(names, start=1):
        ws.cell(row=1, column=col, value=name)

    # 第二行：Fitness
    for col, val in enumerate(fitnesss, start=1):
        ws.cell(row=2, column=col, value=val)

    # 第三行：Precision
    for col, val in enumerate(precisions, start=1):
        ws.cell(row=3, column=col, value=val)

    # 第四行：F-measure
    for col, val in enumerate(f_measures, start=1):
        ws.cell(row=4, column=col, value=val)


    # 行标题（可选，推荐论文用）
    # ws.cell(row=2, column=0 + 1).offset(column=-1, row=0)

    wb.save(OUTPUT_XLSX)


# ================== 入口 ==================

if __name__ == "__main__":
    main()
